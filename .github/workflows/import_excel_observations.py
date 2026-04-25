#!/usr/bin/env python3
"""
Import Excel observations into observations_cli.csv
=====================================================
Reads the "Observations" sheet from the standalone dashboard Excel export
and merges those historical rows into observations_cli.csv.

The old dashboard captured a subset of what the new CLI collector captures.
Columns that weren't in the old format are left blank. The two datasets
join on obs_date so the verification analysis works seamlessly across
the full date range.

Old dashboard Observations columns:
  Date, Max Temp, Min Temp, Avg Temp, Departure, Precip, Snow,
  Avg Wind, Max Wind, Peak Gust, Sky Cover, WX

Mapping notes:
  - Sky Cover: old dashboard stores tenths (0-9) → convert to tenths
    for sky_cover_tenths column (same scale as CLI)
  - Departure: old dashboard only has combined departure, not hi/lo split
    → stored in dep_avg_f; dep_max_f and dep_min_f left blank
  - WX code '8': CF6 WX code 8 = "Smoke, haze" → stored in weather_conditions
  - Wind: old dashboard has avg, max, and peak gust — all map cleanly
  - Missing: normals, record temps, RH, sunshine, precip totals,
    wind direction/timing — all left blank

Usage:
    python import_excel_observations.py \
        --excel forecast-verification-all-2026-04-25.xlsx \
        --csv data/observations_cli.csv

Safe to re-run — deduplicates on obs_date.
"""

import csv
import sys
import os
import argparse
import shutil

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# ── CF6 WX codes → readable descriptions ────────────────────────────────────
CF6_WX_CODES = {
    "1":  "Fog or mist",
    "2":  "Heavy fog",
    "3":  "Thunder",
    "4":  "Ice pellets",
    "5":  "Hail",
    "6":  "Freezing rain or drizzle",
    "7":  "Duststorm or sandstorm",
    "8":  "Smoke or haze",
    "9":  "Blowing snow",
    "X":  "Mixed precipitation",
}

CSV_COLUMNS = [
    "obs_date",
    "product_id",
    "issued_at",
    "retrieved_at",
    "max_temp_f",
    "max_temp_time",
    "min_temp_f",
    "min_temp_time",
    "avg_temp_f",
    "normal_high_f",
    "normal_low_f",
    "normal_avg_f",
    "dep_max_f",
    "dep_min_f",
    "dep_avg_f",
    "record_high_f",
    "record_high_year",
    "record_low_f",
    "record_low_year",
    "precip_today_in",
    "precip_month_in",
    "precip_normal_month_in",
    "precip_year_in",
    "precip_normal_year_in",
    "precip_water_year_in",
    "precip_normal_water_year_in",
    "avg_wind_speed_mph",
    "resultant_wind_speed_mph",
    "resultant_wind_dir_deg",
    "resultant_wind_dir_compass",
    "max_wind_speed_mph",
    "max_wind_dir_deg",
    "max_wind_dir_compass",
    "max_wind_time",
    "peak_gust_mph",
    "peak_gust_dir_deg",
    "peak_gust_dir_compass",
    "peak_gust_time",
    "sky_cover_tenths",
    "sunshine_pct",
    "sunshine_hours",
    "rh_max_pct",
    "rh_max_time",
    "rh_min_pct",
    "rh_min_time",
    "rh_avg_pct",
    "weather_conditions",
]


def parse_precip(val):
    """Handle numeric, T (trace), and 0 precip values."""
    if val is None:
        return ""
    s = str(val).strip().upper()
    if s == "T":
        return "0.001"
    if s in ("0", "0.0", ""):
        return "0.0"
    try:
        return str(round(float(s), 3))
    except ValueError:
        return ""


def parse_wx(val):
    """Convert CF6 WX code to readable description."""
    if val is None:
        return ""
    s = str(val).strip()
    if not s:
        return ""
    # May be multiple codes like "18" or "38"
    descriptions = []
    for ch in s:
        if ch in CF6_WX_CODES:
            descriptions.append(CF6_WX_CODES[ch])
    return ". ".join(descriptions) if descriptions else f"WX code {s}"


def parse_excel_row(row):
    """
    Convert one Observations Excel row to CLI CSV dict.

    Excel columns (0-indexed):
      0  Date
      1  Max Temp
      2  Min Temp
      3  Avg Temp
      4  Departure   (from normal avg — only combined value available)
      5  Precip
      6  Snow
      7  Avg Wind
      8  Max Wind
      9  Peak Gust
      10 Sky Cover   (tenths, 0-9)
      11 WX          (CF6 code string)
    """
    def v(val):
        if val is None:
            return ""
        if isinstance(val, float) and val == int(val):
            return str(int(val))
        return str(val)

    date_val = str(row[0]) if row[0] else ""

    # Sky cover: old dashboard in tenths (0-9) — keep as tenths for CLI schema
    sky_raw = row[10]
    sky_tenths = ""
    if sky_raw is not None and sky_raw != "":
        try:
            sky_tenths = str(round(float(sky_raw), 1))
        except (ValueError, TypeError):
            sky_tenths = ""

    return {
        "obs_date":                    date_val,
        "product_id":                  "IMPORTED-EXCEL",
        "issued_at":                   "",   # not available in old dashboard
        "retrieved_at":                "",   # not available in old dashboard
        "max_temp_f":                  v(row[1]),
        "max_temp_time":               "",   # not captured
        "min_temp_f":                  v(row[2]),
        "min_temp_time":               "",   # not captured
        "avg_temp_f":                  v(row[3]),
        "normal_high_f":               "",   # not in old dashboard
        "normal_low_f":                "",
        "normal_avg_f":                "",
        "dep_max_f":                   "",   # only combined departure available
        "dep_min_f":                   "",
        "dep_avg_f":                   v(row[4]),   # combined departure → avg
        "record_high_f":               "",   # not in old dashboard
        "record_high_year":            "",
        "record_low_f":                "",
        "record_low_year":             "",
        "precip_today_in":             parse_precip(row[5]),
        "precip_month_in":             "",   # not captured
        "precip_normal_month_in":      "",
        "precip_year_in":              "",
        "precip_normal_year_in":       "",
        "precip_water_year_in":        "",
        "precip_normal_water_year_in": "",
        "avg_wind_speed_mph":          v(row[7]),
        "resultant_wind_speed_mph":    "",   # not captured
        "resultant_wind_dir_deg":      "",
        "resultant_wind_dir_compass":  "",
        "max_wind_speed_mph":          v(row[8]),
        "max_wind_dir_deg":            "",   # not captured
        "max_wind_dir_compass":        "",
        "max_wind_time":               "",
        "peak_gust_mph":               v(row[9]),
        "peak_gust_dir_deg":           "",   # not captured
        "peak_gust_dir_compass":       "",
        "peak_gust_time":              "",
        "sky_cover_tenths":            sky_tenths,
        "sunshine_pct":                "",   # not in old dashboard
        "sunshine_hours":              "",
        "rh_max_pct":                  "",   # not in old dashboard
        "rh_max_time":                 "",
        "rh_min_pct":                  "",
        "rh_min_time":                 "",
        "rh_avg_pct":                  "",
        "weather_conditions":          parse_wx(row[11]),
    }


def load_existing_dates(csv_path):
    """Return set of obs_date values already in the CSV."""
    dates = set()
    if not os.path.exists(csv_path):
        return dates
    with open(csv_path, newline="") as f:
        for row in csv.DictReader(f):
            d = row.get("obs_date", "").strip()
            if d:
                dates.add(d)
    return dates


def load_existing_rows(csv_path):
    if not os.path.exists(csv_path):
        return []
    with open(csv_path, newline="") as f:
        return list(csv.DictReader(f))


def main():
    parser = argparse.ArgumentParser(
        description="Import Excel observations into observations_cli.csv"
    )
    parser.add_argument("--excel",   required=True, help="Path to Excel file")
    parser.add_argument("--csv",     required=True, help="Path to observations_cli.csv")
    parser.add_argument("--dry-run", action="store_true", help="Preview only, don't write")
    args = parser.parse_args()

    print(f"Import: Excel observations → {args.csv}")
    print(f"Source: {args.excel}")
    if args.dry_run:
        print("DRY RUN — no files will be modified")

    # Load Excel
    print("\nLoading Excel workbook…")
    wb = load_workbook(args.excel, read_only=True)
    if "Observations" not in wb.sheetnames:
        print("ERROR: 'Observations' sheet not found in workbook")
        sys.exit(1)

    ws = wb["Observations"]
    xl_rows = list(ws.iter_rows(min_row=2, values_only=True))
    print(f"  Found {len(xl_rows)} rows in 'Observations'")

    # Load existing CSV
    existing_dates = load_existing_dates(args.csv)
    existing_rows  = load_existing_rows(args.csv)
    print(f"  Existing CSV rows  : {len(existing_rows)}")
    print(f"  Existing obs dates : {len(existing_dates)}")

    # Parse and deduplicate
    new_rows = []
    skipped  = 0
    errors   = 0

    for i, row in enumerate(xl_rows):
        try:
            parsed = parse_excel_row(row)
        except Exception as e:
            print(f"  WARNING row {i+2}: parse error — {e}")
            errors += 1
            continue

        date = parsed.get("obs_date", "")
        if not date:
            print(f"  WARNING row {i+2}: missing date, skipping")
            errors += 1
            continue

        if date in existing_dates:
            skipped += 1
            continue

        new_rows.append(parsed)
        existing_dates.add(date)

    print(f"\n  New rows to add  : {len(new_rows)}")
    print(f"  Skipped (dups)   : {skipped}")
    print(f"  Parse errors     : {errors}")

    if not new_rows:
        print("\nNothing to add — CSV is already up to date.")
        return

    new_rows.sort(key=lambda r: r["obs_date"])
    print(f"\nDate range of new rows: "
          f"{new_rows[0]['obs_date']} → {new_rows[-1]['obs_date']}")

    # Show sample
    print("\nSample rows:")
    for r in new_rows[:4]:
        print(f"  {r['obs_date']}  Hi={r['max_temp_f']}  Lo={r['min_temp_f']}"
              f"  Wind={r['max_wind_speed_mph']}  Gust={r['peak_gust_mph']}"
              f"  Sky={r['sky_cover_tenths']}/10"
              f"  Precip={r['precip_today_in']}"
              f"  WX={r['weather_conditions'] or '—'}")

    if args.dry_run:
        print("\nDRY RUN complete — no files written.")
        return

    # Merge and sort all rows chronologically
    all_rows = new_rows + [dict(r) for r in existing_rows]
    all_rows.sort(key=lambda r: r.get("obs_date", ""))

    # Backup existing file
    if os.path.exists(args.csv):
        backup_path = args.csv + ".bak"
        shutil.copy2(args.csv, backup_path)
        print(f"\nBackup written to: {backup_path}")

    os.makedirs(os.path.dirname(args.csv) if os.path.dirname(args.csv) else ".", exist_ok=True)

    with open(args.csv, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        for row in all_rows:
            for col in CSV_COLUMNS:
                if col not in row:
                    row[col] = ""
            writer.writerow(row)

    print(f"\nWrote {len(all_rows)} total rows to {args.csv}")
    print(f"  ({len(new_rows)} new historical + {len(existing_rows)} existing)")
    print("\nDone. Review the data then delete the .bak file when satisfied.")
    print("\nNOTE: These imported rows have product_id='IMPORTED-EXCEL'.")
    print("Blank fields (normals, RH, sunshine, wind direction) reflect")
    print("data the old dashboard did not capture — not errors.")


if __name__ == "__main__":
    main()
