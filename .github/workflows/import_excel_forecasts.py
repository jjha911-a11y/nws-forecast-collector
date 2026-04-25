#!/usr/bin/env python3
"""
Import Excel forecast snapshots into forecasts.csv
====================================================
Reads the "All Forecasts" sheet from the standalone dashboard Excel export
and merges those historical rows into the current forecasts.csv, filling
empty values for new fields the old dashboard didn't capture.

Usage:
    python import_excel_forecasts.py \
        --excel forecast-verification-all-2026-04-25.xlsx \
        --csv data/forecasts.csv

Run this once locally or via a one-time GitHub Actions job.
The script is safe to re-run — it deduplicates on
(forecast_updated_at, forecast_date) so no row is written twice.
"""

import csv
import sys
import os
import argparse
from datetime import datetime, timezone

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# ── Column definitions ────────────────────────────────────────────────────────
CSV_COLUMNS = [
    "forecast_updated_at",
    "retrieved_at",
    "forecast_date",
    "lead_days",
    "high_temp_f",
    "low_temp_f",
    "apparent_temp_max_f",    # not in old dashboard — will be blank
    "apparent_temp_min_f",    # not in old dashboard — will be blank
    "max_wind_speed_mph",
    "max_wind_gust_mph",
    "wind_direction",
    "avg_dewpoint_f",          # not in old dashboard — will be blank
    "avg_relative_humidity",   # not in old dashboard — will be blank
    "precip_prob_pct",
    "precip_amount_in",
    "prob_thunder_pct",        # not in old dashboard — will be blank
    "sky_cover_pct",
    "min_visibility_miles",    # not in old dashboard — will be blank
    "day_short_forecast",
    "night_short_forecast",
    "weather_summary",         # not in old dashboard — will be blank
]

# Excel column positions (0-indexed) in "All Forecasts" sheet:
# 0  Forecast Date
# 1  Lead Days
# 2  High Temp (F)
# 3  Low Temp (F)
# 4  Max Wind (mph)
# 5  Max Gust (mph)
# 6  Precip Prob (%)
# 7  Precip Amount (in)
# 8  Sky Cover (%)
# 9  Day Conditions
# 10 Night Conditions
# 11 Wind Direction
# 12 Fetched At
# 13 NWS Updated At
# 14 Snapshot ID


def normalize_timestamp(ts):
    """Normalize various ISO timestamp formats to YYYY-MM-DDTHH:MM:SSZ."""
    if not ts:
        return ""
    s = str(ts).strip()
    # Already in target format
    if s.endswith("Z") and "." not in s:
        return s
    # Has milliseconds: 2026-03-20T04:12:14.679Z → strip ms
    if s.endswith("Z") and "." in s:
        return s.split(".")[0] + "Z"
    # Has +00:00 offset
    if "+00:00" in s:
        return s.replace("+00:00", "").split(".")[0] + "Z"
    # Has .000Z
    return s.split(".")[0] + "Z"


def load_existing_keys(csv_path):
    """Return set of (forecast_updated_at, forecast_date) already in CSV."""
    keys = set()
    if not os.path.exists(csv_path):
        return keys
    with open(csv_path, newline="") as f:
        for row in csv.DictReader(f):
            uat = row.get("forecast_updated_at", "").strip()
            fd  = row.get("forecast_date", "").strip()
            if uat and fd:
                keys.add((uat, fd))
    return keys


def load_existing_rows(csv_path):
    """Return all existing rows as list of dicts."""
    if not os.path.exists(csv_path):
        return []
    with open(csv_path, newline="") as f:
        return list(csv.DictReader(f))


def parse_excel_row(row):
    """Convert one Excel row tuple to a CSV dict."""
    def v(val):
        if val is None or val == "":
            return ""
        if isinstance(val, float) and val == int(val):
            return str(int(val))
        return str(val)

    # Sky cover: old dashboard stored as percent already (0-100)
    sky_pct = ""
    if row[8] is not None and row[8] != "":
        try:
            sky_pct = str(int(round(float(row[8]))))
        except (ValueError, TypeError):
            sky_pct = ""

    # Precip amount: handle T (trace) → 0.001
    precip_raw = row[7]
    if str(precip_raw).upper() == "T":
        precip_val = "0.001"
    else:
        precip_val = v(precip_raw)

    return {
        "forecast_updated_at":   normalize_timestamp(row[13]),
        "retrieved_at":          normalize_timestamp(row[12]),
        "forecast_date":         str(row[0]) if row[0] else "",
        "lead_days":             v(row[1]),
        "high_temp_f":           v(row[2]),
        "low_temp_f":            v(row[3]),
        "apparent_temp_max_f":   "",   # not captured by old dashboard
        "apparent_temp_min_f":   "",
        "max_wind_speed_mph":    v(row[4]),
        "max_wind_gust_mph":     v(row[5]),
        "wind_direction":        v(row[11]),
        "avg_dewpoint_f":        "",   # not captured by old dashboard
        "avg_relative_humidity": "",
        "precip_prob_pct":       v(row[6]),
        "precip_amount_in":      precip_val,
        "prob_thunder_pct":      "",   # not captured by old dashboard
        "sky_cover_pct":         sky_pct,
        "min_visibility_miles":  "",   # not captured by old dashboard
        "day_short_forecast":    v(row[9]),
        "night_short_forecast":  v(row[10]),
        "weather_summary":       "",   # not captured by old dashboard
    }


def main():
    parser = argparse.ArgumentParser(description="Import Excel forecasts into forecasts.csv")
    parser.add_argument("--excel", required=True, help="Path to Excel file")
    parser.add_argument("--csv",   required=True, help="Path to forecasts.csv")
    parser.add_argument("--dry-run", action="store_true", help="Preview only, don't write")
    args = parser.parse_args()

    print(f"Import: Excel forecasts → {args.csv}")
    print(f"Source: {args.excel}")
    if args.dry_run:
        print("DRY RUN — no files will be modified")

    # Load Excel
    print("\nLoading Excel workbook…")
    wb = load_workbook(args.excel, read_only=True)
    if "All Forecasts" not in wb.sheetnames:
        print("ERROR: 'All Forecasts' sheet not found in workbook")
        sys.exit(1)

    ws = wb["All Forecasts"]
    xl_rows = list(ws.iter_rows(min_row=2, values_only=True))
    print(f"  Found {len(xl_rows)} rows in 'All Forecasts'")

    # Load existing CSV
    existing_keys = load_existing_keys(args.csv)
    existing_rows = load_existing_rows(args.csv)
    print(f"  Existing CSV rows: {len(existing_rows)}")
    print(f"  Existing dedup keys: {len(existing_keys)}")

    # Parse and deduplicate Excel rows
    new_rows = []
    skipped  = 0
    errors   = 0

    for i, row in enumerate(xl_rows):
        try:
            parsed = parse_excel_row(row)
        except Exception as e:
            print(f"  WARNING row {i+2}: parse error — {e}")
            errors += 1
            continue

        key = (parsed["forecast_updated_at"], parsed["forecast_date"])
        if not key[0] or not key[1]:
            print(f"  WARNING row {i+2}: missing timestamp or date, skipping")
            errors += 1
            continue

        if key in existing_keys:
            skipped += 1
            continue

        new_rows.append(parsed)
        existing_keys.add(key)

    print(f"\n  New rows to add  : {len(new_rows)}")
    print(f"  Skipped (dups)   : {skipped}")
    print(f"  Parse errors     : {errors}")

    if not new_rows:
        print("\nNothing to add — CSV is already up to date.")
        return

    # Sort new rows chronologically by (forecast_updated_at, forecast_date)
    new_rows.sort(key=lambda r: (r["forecast_updated_at"], r["forecast_date"]))

    # Preview
    print(f"\nDate range of new rows: "
          f"{new_rows[0]['forecast_date']} → {new_rows[-1]['forecast_date']}")
    print(f"Snapshot range: "
          f"{new_rows[0]['forecast_updated_at'][:10]} → {new_rows[-1]['forecast_updated_at'][:10]}")

    if args.dry_run:
        print("\nDRY RUN — sample of rows that would be added:")
        for r in new_rows[:3]:
            print(f"  {r['forecast_date']}  D{r['lead_days']}"
                  f"  Hi={r['high_temp_f']}  Lo={r['low_temp_f']}"
                  f"  Wind={r['max_wind_speed_mph']}"
                  f"  Sky={r['sky_cover_pct']}%"
                  f"  [{r['forecast_updated_at'][:16]}]")
        return

    # Rebuild CSV: new historical rows first, then existing rows
    # (ensures chronological order)
    os.makedirs(os.path.dirname(args.csv) if os.path.dirname(args.csv) else ".", exist_ok=True)

    # Merge: put excel rows before existing rows (they're older)
    # but only if they're actually older — sort everything together
    all_rows = new_rows + [dict(r) for r in existing_rows]
    all_rows.sort(key=lambda r: (
        r.get("forecast_updated_at", ""),
        r.get("forecast_date", ""),
    ))

    # Write out
    backup_path = args.csv + ".bak"
    if os.path.exists(args.csv):
        import shutil
        shutil.copy2(args.csv, backup_path)
        print(f"\nBackup written to: {backup_path}")

    with open(args.csv, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        for row in all_rows:
            # Ensure all columns exist (old rows may lack new fields)
            for col in CSV_COLUMNS:
                if col not in row:
                    row[col] = ""
            writer.writerow(row)

    print(f"\nWrote {len(all_rows)} total rows to {args.csv}")
    print(f"  ({len(new_rows)} new historical + {len(existing_rows)} existing)")
    print("\nDone. Review the data then delete the .bak file when satisfied.")


if __name__ == "__main__":
    main()
